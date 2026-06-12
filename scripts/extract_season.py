# -*- coding: utf-8 -*-
"""
Fase 1 de la migración de la temporada 25/26.
Lee el xlsx exportado de Google Sheets y genera:
  - migration/seed.json   (tortillas + votos + referencia de imagen)
  - migration/images/     (imagen principal de cada tortilla)

El matching master<->hoja se hace por fecha (del nombre de hoja) y, si hay
varias tortillas el mismo día, por cercanía entre la media de los votos de la
hoja y la nota apuntada en el master.

Casos especiales que el matching automático no cubre (verificados por media):
  - "Bacon Jam & Queso Brie" (sin fecha en master)  -> hoja sem250924
  - "Tortischorra" (master 8/10, hoja 15/10)        -> manda la hoja
  - "Torrezno..." (master 25/03, hoja sem-260408)   -> manda la hoja
"""
import json
import os
import re
import sys
import zipfile
from datetime import date
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

XLSX = r'C:\Users\Pablo\Downloads\Tortillas Temporada 25_26.xlsx'
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'migration')

# Las pestañas 00* son "fuera de concurso" (sin fecha ni votos): no se migran.
INCLUDE_OOF = False

# Unificación de votantes. Solo typos/apodos evidentes verificados en las hojas.
# OJO: "Pablo" vs "Pako" vs "PabloJuan" NO se unifican aquí — si son la misma
# persona, añade la línea correspondiente antes de ejecutar.
VOTER_ALIASES = {
    'Oswlado': 'Oswaldo',
    'Osw': 'Oswaldo',
    'Victortilla': 'Victor',
    'Victor Tortilla': 'Victor',
}

NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'm': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
}


def parse_sheet_date(sheet_name):
    """sem250903 / sem-260121-xis -> date. None si no es hoja semanal."""
    m = re.match(r'^sem-?(\d{2})(\d{2})(\d{2})', sheet_name)
    if not m:
        return None
    yy, mm, dd = (int(g) for g in m.groups())
    return date(2000 + yy, mm, dd)


def parse_master_date(value):
    if isinstance(value, date):
        return date(value.year, value.month, value.day)
    if isinstance(value, str):
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', value.strip())
        if m:
            d, mth, y = (int(g) for g in m.groups())
            return date(y, mth, d)
    return None


def parse_score(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def sheet_images(z, sheet_targets):
    """Para cada hoja, lista [(anchor_row, media_filename)] ordenada por fila."""
    result = {}
    for sheet_name, target in sheet_targets.items():
        rels_path = 'xl/' + target.replace('worksheets/', 'worksheets/_rels/') + '.rels'
        images = []
        try:
            rels = ET.fromstring(z.read(rels_path))
        except KeyError:
            result[sheet_name] = images
            continue
        for rel in rels:
            if 'drawing' not in rel.get('Type'):
                continue
            drawing_path = 'xl/' + rel.get('Target').replace('../', '')
            drels_path = drawing_path.replace('drawings/', 'drawings/_rels/') + '.rels'
            try:
                drels = ET.fromstring(z.read(drels_path))
                drawing = ET.fromstring(z.read(drawing_path))
            except KeyError:
                continue
            rid_to_media = {
                dr.get('Id'): dr.get('Target').replace('../media/', '')
                for dr in drels
                if 'image' in dr.get('Type')
            }
            for anchor in drawing:
                frm = anchor.find('xdr:from', NS)
                row = int(frm.find('xdr:row', NS).text) if frm is not None else 0
                blip = anchor.find('.//a:blip', NS)
                if blip is None:
                    continue
                rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                media = rid_to_media.get(rid)
                if media:
                    images.append((row, media))
        images.sort(key=lambda t: t[0])
        result[sheet_name] = images
    return result


def main():
    wb = load_workbook(XLSX, data_only=True)
    z = zipfile.ZipFile(XLSX)

    # --- mapping hoja -> sheetN.xml (vía workbook.xml + rels) ---
    wb_root = ET.fromstring(z.read('xl/workbook.xml'))
    wb_rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    rid_to_target = {r.get('Id'): r.get('Target') for r in wb_rels}
    sheet_targets = {}
    for s in wb_root.find('m:sheets', NS):
        rid = s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        sheet_targets[s.get('name')] = rid_to_target[rid]

    images_by_sheet = sheet_images(z, sheet_targets)

    # --- master ---
    master_rows = []
    ws = wb['master']
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name or not str(name).strip():
            continue
        is_oof = len(row) > 5 and str(row[5] or '').strip().upper() == 'X'
        master_rows.append({
            'date': parse_master_date(row[0]),
            'name': str(name).strip(),
            'ingredients': str(row[2]).strip() if row[2] else None,
            'score': parse_score(row[3]),
            'oof': is_oof,
            'note': str(row[6]).strip() if len(row) > 6 and row[6] else None,
        })

    # --- hojas semanales: votos ---
    week_sheets = {}
    for sheet_name in wb.sheetnames:
        d = parse_sheet_date(sheet_name)
        if d is None:
            continue
        ws = wb[sheet_name]
        votes = []
        for row in ws.iter_rows(values_only=True):
            voter = row[0]
            if voter is None or not str(voter).strip():
                continue
            score = parse_score(row[1] if len(row) > 1 else None)
            if score is None:
                continue  # "NO VOTA" o celda vacía
            raw_name = str(voter).strip()
            user_name = VOTER_ALIASES.get(raw_name, raw_name)
            votes.append({'userName': user_name, 'score': round(score, 1)})
        avg = round(sum(v['score'] for v in votes) / len(votes), 2) if votes else None
        week_sheets[sheet_name] = {'date': d, 'votes': votes, 'avg': avg}

    # --- matching ---
    # Casos especiales: hoja -> nombre exacto de fila master.
    SPECIAL = {
        'sem250924': 'Bacon Jam & Queso Brie',
        'sem251015': 'Tortischorra',
        'sem-260408': 'Torrezno i postre Guillem Tiramixu cremaet',
    }

    unmatched_master = [r for r in master_rows if not r['oof']]
    tortillas = []
    warnings = []

    for sheet_name, info in sorted(week_sheets.items(), key=lambda kv: kv[1]['date']):
        if sheet_name in SPECIAL:
            target_name = SPECIAL[sheet_name]
            candidates = [r for r in unmatched_master if r['name'] == target_name]
        else:
            candidates = [r for r in unmatched_master if r['date'] == info['date']]

        if not candidates:
            warnings.append(f'{sheet_name}: sin fila en master — se omite')
            continue

        if len(candidates) == 1:
            chosen = candidates[0]
        else:
            # Varias tortillas el mismo día: matching por nota media.
            chosen = min(
                candidates,
                key=lambda r: abs((r['score'] or 0) - (info['avg'] or 0)),
            )

        unmatched_master.remove(chosen)
        diff = (
            abs((chosen['score'] or 0) - (info['avg'] or 0))
            if chosen['score'] is not None and info['avg'] is not None
            else None
        )
        if diff is not None and diff > 0.05:
            warnings.append(
                f'{sheet_name} -> "{chosen["name"]}": media hoja {info["avg"]} '
                f'vs master {chosen["score"]} (diff {diff:.2f})'
            )
        if chosen['date'] and chosen['date'] != info['date']:
            warnings.append(
                f'{sheet_name} -> "{chosen["name"]}": fecha hoja {info["date"]} '
                f'!= master {chosen["date"]} (manda la hoja)'
            )

        imgs = images_by_sheet.get(sheet_name, [])
        tortillas.append({
            'sheet': sheet_name,
            'name': chosen['name'],
            'description': chosen['ingredients'],
            'date': info['date'].isoformat(),
            'votes': info['votes'],
            'masterScore': chosen['score'],
            'sheetAvg': info['avg'],
            'image': imgs[0][1] if imgs else None,
            'extraImages': [m for _, m in imgs[1:]],
        })

    # Filas master de temporada que quedaron sin hoja (sin votos ni imagen).
    for r in unmatched_master:
        if r['date'] is None:
            warnings.append(f'master "{r["name"]}": sin fecha y sin hoja — se omite')
            continue
        warnings.append(
            f'master "{r["name"]}" ({r["date"]}): sin hoja de votos — '
            f'se migra sin votos individuales (media quedará vacía) y sin foto'
        )
        tortillas.append({
            'sheet': None,
            'name': r['name'],
            'description': r['ingredients'],
            'date': r['date'].isoformat(),
            'votes': [],
            'masterScore': r['score'],
            'sheetAvg': None,
            'image': None,
            'extraImages': [],
        })

    tortillas.sort(key=lambda t: t['date'])

    # --- extraer imágenes usadas ---
    img_dir = os.path.join(OUT_DIR, 'images')
    os.makedirs(img_dir, exist_ok=True)
    used = set()
    for t in tortillas:
        for media in ([t['image']] if t['image'] else []) + t['extraImages']:
            used.add(media)
    for media in sorted(used):
        with open(os.path.join(img_dir, media), 'wb') as f:
            f.write(z.read(f'xl/media/{media}'))

    # Placeholder para tortillas sin foto (imageKey es obligatorio en el modelo).
    needs_placeholder = [t for t in tortillas if not t['image']]
    if needs_placeholder:
        try:
            from PIL import Image
        except ImportError:
            print('AVISO: falta Pillow (pip install pillow) para generar el placeholder.')
            Image = None
        if Image:
            img = Image.new('RGB', (1200, 900), '#ffeccb')
            img.save(os.path.join(img_dir, 'placeholder.jpg'), quality=85)
            for t in needs_placeholder:
                t['image'] = 'placeholder.jpg'

    seed = {'season': '2025-2026', 'tortillas': tortillas}
    with open(os.path.join(OUT_DIR, 'seed.json'), 'w', encoding='utf-8') as f:
        json.dump(seed, f, ensure_ascii=False, indent=2)

    # --- reporte ---
    print(f'Tortillas: {len(tortillas)}')
    total_votes = sum(len(t['votes']) for t in tortillas)
    print(f'Votos individuales: {total_votes}')
    print(f'Imagenes extraidas: {len(used)} -> {img_dir}')
    voters = sorted({v['userName'] for t in tortillas for v in t['votes']})
    print(f'Votantes unicos ({len(voters)}): {", ".join(voters)}')
    print()
    print('Plan (fecha | nombre | votos | media hoja vs master | imagen):')
    for t in tortillas:
        img = t['image'] or 'SIN IMAGEN'
        print(
            f"  {t['date']} | {t['name'][:45]:45} | {len(t['votes']):2} votos | "
            f"{str(t['sheetAvg']):5} vs {str(t['masterScore']):5} | {img}"
        )
    if warnings:
        print()
        print('AVISOS:')
        for w in warnings:
            print(f'  - {w}')


if __name__ == '__main__':
    sys.exit(main())
